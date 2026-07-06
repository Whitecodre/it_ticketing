import random, hashlib, os, re, csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.management import call_command
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Avg, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.utils.html import strip_tags
from django.urls import reverse
from django.template.loader import render_to_string
from datetime import timedelta, datetime
from django.core.mail import send_mail
from django.conf import settings
from .forms import TicketForm, CommentForm, AssetForm
from .models import *
from apps.accounts.models import User
from apps.common.models import Notification
from bs4 import BeautifulSoup

import logging
logger = logging.getLogger(__name__)

# ==========================================================================
# HELPER FUNCTIONS
# ==========================================================================

def clean_comment_body(body):
    """
    Cleans HTML content from the rich text editor before saving.
    - Removes empty div/p/span tags.
    - Replaces <br> tags with newlines.
    - Collapses multiple consecutive newlines into a single newline.
    - Strips leading/trailing whitespace.
    Returns None if the cleaned body is empty (used to reject empty comments).
    """
    if not body:
        return None
    soup = BeautifulSoup(body, 'html.parser')
    
    # Remove empty block tags that contain no text and no formatting children
    for tag in soup.find_all():
        if tag.name in ['div', 'p', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
            if not tag.get_text(strip=True) and not tag.find_all(['strong', 'em', 'br']):
                tag.decompose()
    
    # Replace <br> with newline characters (they will be turned back into <br> by linebreaksbr filter)
    for br in soup.find_all('br'):
        br.replace_with('\n')
    
    cleaned = str(soup)
    cleaned = re.sub(r'\n\s*\n+', '\n', cleaned)
    cleaned = cleaned.strip()
    
    if not cleaned or cleaned == '':
        return None
    return cleaned

def get_sidebar_template(user):
    """
    Returns the correct sidebar partial template based on the user's role.
    Used in all dashboard views to load the appropriate navigation sidebar.
    """
    mapping = {
        'END_USER': 'partials/sidebar_end_user.html',
        'AGENT': 'partials/sidebar_agent.html',
        'TEAM_LEAD': 'partials/sidebar_team_lead.html',
        'ADMIN': 'partials/sidebar_admin.html',
        'SUPERADMIN': 'partials/sidebar_superadmin.html',
        # 'APPROVER': 'partials/sidebar_approver.html',
    }
    return mapping.get(user.role, 'partials/sidebar_end_user.html')

def apply_sla(ticket):
    """
    Sets the response_due_at and resolution_due_at fields on a ticket
    based on the SLA policy configured for its priority.
    Called after ticket creation and when priority changes.
    """
    try:
        sla = SLA.objects.get(priority=ticket.priority)
    except SLA.DoesNotExist:
        return
    now = timezone.now()
    ticket.response_due_at = now + timedelta(minutes=sla.response_minutes)
    ticket.resolution_due_at = now + timedelta(minutes=sla.resolution_minutes)
    ticket.save(update_fields=['response_due_at', 'resolution_due_at'])

# Helper function to handle "Other" field logic
def get_other_value(data, select_field, other_field, default_value):
    """Helper to handle 'Other' field logic for asset forms."""
    value = data.get(select_field)
    if value == 'OTHER':
        other_val = data.get(other_field, '').strip()
        return other_val if other_val else default_value
    return value

# Allowed MIME types for file attachments
ALLOWED_MIMES = [
    'image/jpeg', 'image/png', 'image/gif', 'image/webp',
    'application/pdf',
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'text/plain',
    'application/zip',
]
MAX_SIZE_MB = 10

def save_attachments(ticket, files, author, comment=None):
    """
    Validates and saves uploaded file attachments.
    - Skips files that exceed MAX_SIZE_MB or have disallowed MIME types.
    - Computes SHA-256 hash for integrity checking.
    - Associates attachments with a ticket and optionally a specific comment.
    Returns list of created Attachment objects.
    """
    created = []
    for f in files:
        if f.size > MAX_SIZE_MB * 1024 * 1024:
            continue
        mime = f.content_type.split(';')[0].strip().lower()
        if mime not in ALLOWED_MIMES:
            continue
        sha = hashlib.sha256()
        for chunk in f.chunks():
            sha.update(chunk)
        f.seek(0)
        att = Attachment.objects.create(
            ticket=ticket,
            comment=comment,
            file=f,
            filename=f.name,
            uploaded_by=author,
            content_type=mime,
            size=f.size,
            hash=sha.hexdigest(),
        )
        created.append(att)
    return created

# ==========================================================================
# TICKET CREATION & LISTING VIEWS (End Users)
# ==========================================================================

# apps/tickets/views.py
import random
from django.contrib import messages

@login_required
def create_ticket(request):
    """
    Handles creation of a new ticket (incident or service request).
    - GET: displays the appropriate form (incident_form or service_request_form).
    - POST: validates form, generates a unique ticket number, saves ticket,
            attaches files, applies SLA, and redirects to ticket detail page.
    """
    ticket_type = request.GET.get('type', 'INCIDENT').upper()
    if ticket_type not in ['INCIDENT', 'SERVICE_REQUEST']:
        ticket_type = 'INCIDENT'

    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.requester = request.user
            ticket.type = ticket_type

            # Generate unique ticket number (e.g., TK#1234 or SRV#5678)
            prefix = 'TK' if ticket.type == Ticket.Type.INCIDENT else 'SRV'
            for _ in range(20):
                suffix = str(random.randint(0, 9999)).zfill(4)
                candidate = f"{prefix}#{suffix}"
                if not Ticket.objects.filter(number=candidate).exists():
                    ticket.number = candidate
                    break
            else:
                import time
                ticket.number = f"{prefix}#{int(time.time()) % 10000:04d}"

            ticket.save()

            files = request.FILES.getlist('attachments')
            if files:
                save_attachments(ticket, files, request.user)

            # If it's a service request, set status to PENDING_MANAGER_REVIEW
            if ticket.type == Ticket.Type.SERVICE_REQUEST:
                ticket.status = Ticket.Status.PENDING_MANAGER_REVIEW

                # Check if this is an asset-related request
                asset_categories = ['Hardware', 'Item acquisition', 'Software', 'Purchase/Protocol']
                if ticket.category and ticket.category.name in asset_categories:
                    ticket.is_asset_request = True
                    print(f"🔵 is_asset_request set to True for category: {ticket.category.name}")

                # FIX: Save BOTH status and is_asset_request
                ticket.save(update_fields=['status', 'is_asset_request'])

                
                messages.success(request, f'Service request {ticket.number} submitted for manager review.')
            else:
                messages.success(request, f'Ticket {ticket.number} created successfully.')

            apply_sla(ticket)

            return redirect('tickets:detail', pk=ticket.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = TicketForm(initial={'type': ticket_type})

    template = 'requester/incident_form.html' if ticket_type == 'INCIDENT' else 'requester/service_request_form.html'
    return render(request, template, {'form': form, 'ticket_type': ticket_type})

@login_required
@require_POST
def cancel_ticket(request, pk):
    """
    Allows an end user to cancel a ticket that is still in NEW or TRIAGED status.
    - Closes the ticket and logs the action.
    - If the request is HTMX, returns the updated ticket list partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester:
        return HttpResponse(status=403)
    if ticket.status not in [Ticket.Status.NEW, Ticket.Status.TRIAGED]:
        return HttpResponse(status=400)
    ticket.status = Ticket.Status.CLOSED
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket, action='status_changed', actor=request.user,
        details={'from': ticket.status, 'to': Ticket.Status.CLOSED, 'reason': 'Cancelled by requester'}
    )
    if request.headers.get('HX-Request'):
        tickets = Ticket.objects.filter(requester=request.user).order_by('-created_at')
        status_filter = request.POST.get('current_status', '')
        if status_filter and status_filter.upper() in dict(Ticket.Status.choices):
            tickets = tickets.filter(status=status_filter.upper())
        paginator = Paginator(tickets, 10)
        page_number = request.POST.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        context = {
            'tickets': page_obj,
            'current_status': status_filter,
            'status_choices': Ticket.Status.choices,
        }
        return render(request, 'partials/ticket_list_partial.html', context)
    return redirect('tickets:my_list')

@login_required
def my_ticket_list(request):
    """
    Displays a list of tickets created by the logged‑in end user.
    Supports filtering by status (OPEN/CLOSED or specific status).
    Uses HTMX for pagination and filter updates.
    """
    tickets = Ticket.objects.filter(requester=request.user).order_by('-created_at')
    status_filter = request.GET.get('status', '')
    base = request.GET.get('base', '')
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR', 'APPROVED']
    closed_statuses = ['RESOLVED', 'CLOSED']

    if status_filter == 'OPEN':
        tickets = tickets.filter(status__in=open_statuses)
    elif status_filter == 'CLOSED':
        tickets = tickets.filter(status__in=closed_statuses)
    elif status_filter and status_filter.upper() in dict(Ticket.Status.choices):
        tickets = tickets.filter(status=status_filter.upper())

    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    all_choices = Ticket.Status.choices
    if base == 'OPEN':
        status_choices = [('NEW','New'), ('TRIAGED','Triaged'), ('ASSIGNED','Assigned'),
                          ('IN_PROGRESS','In Progress'), ('PENDING_USER','Pending User'),
                          ('PENDING_VENDOR','Pending Vendor'), ('APPROVED','Approved')]
    elif base == 'CLOSED':
        status_choices = [('RESOLVED','Resolved'), ('CLOSED','Closed')]
    else:
        status_choices = all_choices

    base_status = base if base in ['OPEN','CLOSED'] else ''
    explicit = request.GET.get('explicit') == '1'

    context = {
        'tickets': page_obj,
        'current_status': status_filter or '',
        'status_choices': status_choices,
        'sidebar_template': get_sidebar_template(request.user),
        'base_status': base_status,
        'explicit': explicit,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/ticket_list_partial.html', context)
    return render(request, 'requester/ticket_list.html', context)

@login_required
def ticket_detail(request, pk):
    """
    Unified ticket conversation page for both requesters and agents.
    - GET: displays the conversation timeline and comment form.
    - POST (HTMX): accepts a new public comment from the requester,
      cleans the HTML body, saves attachments, updates ticket status if needed,
      and returns the updated timeline partial.
    The 'is_agent' flag controls visibility of agent‑only UI elements.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user != ticket.requester and request.user.role not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return redirect('dashboard')

    # Handle comment submission from requester
    if request.method == 'POST' and request.headers.get('HX-Request'):
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.ticket = ticket
            comment.author = request.user
            comment.visibility = 'PUBLIC'
            cleaned_body = clean_comment_body(comment.body)
            if cleaned_body is None:
                return HttpResponse(status=400)   # empty message
            comment.body = cleaned_body
            comment.save()

            files = request.FILES.getlist('attachments')
            if files:
                save_attachments(ticket, files, request.user, comment=comment)

            if ticket.status == Ticket.Status.PENDING_USER:
                ticket.status = Ticket.Status.IN_PROGRESS
                ticket.save()
                TicketActivityLog.objects.create(
                    ticket=ticket, action='status_changed', actor=request.user,
                    details={'from': Ticket.Status.PENDING_USER, 'to': Ticket.Status.IN_PROGRESS}
                )

            comments = ticket.comments.prefetch_related('attachment_set').all().order_by('created_at')
            initial_attachments = ticket.attachments.filter(comment__isnull=True)
            return render(request, 'partials/conversation_timeline.html', {
                'ticket': ticket,
                'comments': comments,
                'initial_attachments': initial_attachments,
            })
        else:
            return HttpResponse(status=422)

    # GET request – render conversation page
    comments = ticket.comments.all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    form = CommentForm()
    is_agent = request.user.role in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']

    return render(request, 'agent/ticket_conversation.html', {
        'ticket': ticket,
        'comments': comments,
        'form': form,
        'initial_attachments': initial_attachments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
        'sidebar_template': get_sidebar_template(request.user),
        'is_agent': is_agent,
    })

# ==========================================================================
# AGENT QUEUES & TICKET MANAGEMENT
# ==========================================================================

@login_required
def unassigned_queue(request):
    """
    Displays all unassigned tickets that are ready for agents to claim.
    """
    # Get all unassigned tickets
    tickets = Ticket.objects.filter(
        assigned_to__isnull=True
    ).order_by('-created_at')
    
    # ================================================================
    # 🔥 Filter: Show INCIDENTS + APPROVED SERVICE REQUESTS
    # ================================================================
    tickets = tickets.filter(
        Q(type=Ticket.Type.INCIDENT) |
        (Q(type=Ticket.Type.SERVICE_REQUEST) & Q(status=Ticket.Status.APPROVED))
    )
    
    # Exclude tickets that shouldn't be in the queue
    tickets = tickets.exclude(
        status__in=[
            Ticket.Status.PENDING_MANAGER_REVIEW,
            Ticket.Status.PENDING_FULFILLMENT,
            Ticket.Status.PENDING_APPROVAL,
            Ticket.Status.RESOLVED,
            Ticket.Status.CLOSED,
        ]
    )

    assignable_agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    ).only('pk', 'first_name', 'last_name', 'email')

    context = {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'agent/unassigned_queue.html', context)

@login_required
def assigned_to_me(request):
    """
    Displays all tickets assigned to the logged‑in agent,
    excluding resolved, closed, and pending approval tickets.
    """
    tickets = Ticket.objects.filter(
        assigned_to=request.user
    ).exclude(
        status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL, Ticket.Status.PENDING_MANAGER_REVIEW]
    ).order_by('-created_at')

    assignable_agents = User.objects.filter(role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']).only('pk', 'first_name', 'last_name', 'email')

    context = {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'agent/assigned_to_me.html', context)

@login_required
def claim_ticket(request, pk):
    """
    Allows an agent to claim an unassigned ticket.
    Assigns the ticket to the current user and sets status to ASSIGNED.
    Returns the updated agent ticket table partial (HTMX).
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if ticket.assigned_to is None:
        # Prevent claiming if ticket is pending manager review
        if ticket.status == Ticket.Status.PENDING_MANAGER_REVIEW:
            return HttpResponse("This ticket is pending manager review and cannot be claimed.", status=400)
        ticket.assigned_to = request.user
        ticket.status = Ticket.Status.ASSIGNED
        ticket.save()
        TicketActivityLog.objects.create(
            ticket=ticket, action='assigned', actor=request.user,
            details={'to': request.user.get_full_name(), 'status': ticket.status}
        )
    tickets = Ticket.objects.filter(assigned_to__isnull=True).exclude(
        status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL]
    ).order_by('-created_at')

    assignable_agents = User.objects.filter(
        role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    ).only('pk', 'first_name', 'last_name', 'email')

    return render(request, 'partials/agent_ticket_table.html', {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
    })

@login_required
def agent_ticket_detail(request, pk):
    """
    Returns a slide‑over panel with ticket details and comments.
    Used when an agent clicks the "eye" icon on a ticket row.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    comments = ticket.comments.all().order_by('created_at')
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'partials/ticket_slideover.html', {
        'ticket': ticket,
        'comments': comments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
    })

@login_required
def agent_ticket_conversation(request, pk):
    """
    Full‑page conversation view for agents (and admins).
    Renders the same template as ticket_detail but with is_agent=True.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    comments = ticket.comments.all().order_by('created_at')
    form = CommentForm()
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'agent/ticket_conversation.html', {
        'ticket': ticket,
        'comments': comments,
        'form': form,
        'initial_attachments': initial_attachments,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
        'sidebar_template': get_sidebar_template(request.user),
        'is_agent': True,
    })

@login_required
@require_POST
def add_comment_conversation(request, pk):
    """
    Handles agent comments (public or internal) on a ticket.
    - Cleans the HTML body using BeautifulSoup.
    - Saves attachments.
    - Updates ticket status to IN_PROGRESS if a public comment is added.
    - Returns the updated conversation timeline partial (HTMX).
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    form = CommentForm(request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.ticket = ticket
        comment.author = request.user
        comment.visibility = request.POST.get('visibility', 'PUBLIC').upper()
        if comment.visibility not in ['PUBLIC', 'INTERNAL']:
            comment.visibility = 'PUBLIC'
        cleaned_body = clean_comment_body(comment.body)
        if cleaned_body is None:
            return HttpResponse(status=400)
        comment.body = cleaned_body
        comment.save()

        files = request.FILES.getlist('attachments')
        if files:
            created = save_attachments(ticket, files, request.user, comment=comment)
            comment = TicketComment.objects.get(pk=comment.pk)

        TicketActivityLog.objects.create(
            ticket=ticket, action='commented', actor=request.user,
            details={'visibility': comment.visibility, 'body': comment.body[:200]}
        )

        old_status = ticket.status
        if comment.visibility == 'PUBLIC':
            if old_status in [Ticket.Status.ASSIGNED, Ticket.Status.IN_PROGRESS,
                              Ticket.Status.PENDING_USER, Ticket.Status.NEW, Ticket.Status.TRIAGED]:
                ticket.status = Ticket.Status.IN_PROGRESS
                ticket.save()
                if old_status != ticket.status:
                    TicketActivityLog.objects.create(
                        ticket=ticket, action='status_changed', actor=request.user,
                        details={'from': old_status, 'to': ticket.status}
                    )

    comments = ticket.comments.prefetch_related('attachment_set').all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)
    return render(request, 'partials/conversation_timeline.html', {
        'ticket': ticket,
        'comments': comments,
        'initial_attachments': initial_attachments, 
    })

@login_required
@require_POST
def update_status(request, pk):
    """
    Changes the status of a ticket. Only agents, team leads, admins, and superadmins can do this.
    The new status is passed as a GET parameter 'status'.
    Logs the change in TicketActivityLog.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    
    previous_status = ticket.status
    new_status = request.GET.get('status')
    if new_status and new_status in dict(Ticket.Status.choices):
        ticket.status = new_status
        ticket.save()
        TicketActivityLog.objects.create(
            ticket=ticket, action='status_changed', actor=request.user,
            details={'from': previous_status, 'to': new_status}
        )
    return HttpResponse(status=204)

# ==========================================================================
# TICKET DETAILS PANEL & METADATA EDITING
# ==========================================================================

@login_required
def ticket_details_panel(request, pk):
    """
    Returns the right‑hand details panel (assignee, metadata, attachments)
    that slides in on the conversation page.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    followers = User.objects.filter(role__in=['AGENT','TEAM_LEAD'])[:5]
    user_attachments = ticket.attachments.filter(uploaded_by__role='END_USER')
    agent_attachments = ticket.attachments.filter(
        uploaded_by__role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']
    )
    return render(request, 'partials/ticket_details_panel.html', {
        'ticket': ticket,
        'followers': followers,
        'user_attachments': user_attachments,
        'agent_attachments': agent_attachments,
    })

@login_required
@require_POST
def edit_subject(request, pk):
    """
    Edits the ticket title inline (subject).
    Returns the updated subject display partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    new_title = request.POST.get('title', '').strip()
    if new_title:
        ticket.title = new_title
        ticket.save()
    return render(request, 'partials/subject_display.html', {'ticket': ticket})

# ==========================================================================
# ASSIGNMENT POPOVERS AND ACTIONS
# ==========================================================================

@login_required
def assign_popover(request, pk):
    """
    Returns a popover with a list of assignable agents (Agent, Team Lead, Admin, Superadmin)
    so the agent can reassign the ticket.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    agents = User.objects.filter(role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN'])[:10]
    return render(request, 'partials/popovers/assign_popover.html', {'ticket': ticket, 'agents': agents})

@login_required
@require_POST
def assign_to_me(request, pk):
    """
    Assigns the ticket to the current user.
    Returns the updated assignee display partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    ticket.assigned_to = request.user
    ticket.save()
    return render(request, 'partials/ticket_details_assignee.html', {'ticket': ticket})

@login_required
@require_POST
def assign_specific(request, pk, user_pk):
    """
    Assigns the ticket to a specific user (by primary key).
    Returns the updated assignee display partial.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    agent = get_object_or_404(User, pk=user_pk)
    ticket.assigned_to = agent
    ticket.save()
    return render(request, 'partials/ticket_details_assignee.html', {'ticket': ticket})

# ==========================================================================
# FOLLOWER STUBS (not fully implemented)
# ==========================================================================

@login_required
@require_POST
def remove_follower(request, ticket_pk, user_pk):
    return HttpResponse(status=204)

@login_required
def add_follower_popover(request, ticket_pk):
    return render(request, 'partials/popovers/add_follower_popover.html', {
        'ticket': get_object_or_404(Ticket, pk=ticket_pk),
        'agents': User.objects.filter(role__in=['AGENT','TEAM_LEAD'])[:10],
    })

# Placeholder popovers
def edit_group_popover(request, pk): return render(request, 'partials/popovers/empty.html')
def edit_type_popover(request, pk): return render(request, 'partials/popovers/empty.html')
def edit_priority_popover(request, pk): return render(request, 'partials/popovers/empty.html')
def add_tag_popover(request, pk): return render(request, 'partials/popovers/empty.html')
def remove_tag(request, pk, tag_pk): return HttpResponse(status=204)

# ==========================================================================
# MACROS
# ==========================================================================

@login_required
def macro_list(request):
    """
    Returns a dropdown list of macros (predefined reply templates) for agents.
    Used in the conversation composer.
    """
    macros = Macro.objects.all()
    return render(request, 'partials/macro_dropdown.html', {'macros': macros})

# ==========================================================================
# BULK ACTIONS (for ticket queues)
# ==========================================================================

@login_required
@require_POST
def bulk_action(request):
    """
    Performs bulk status change or assignment on multiple tickets.
    Only Team Leads, Admins, and Superadmins can bulk‑assign.
    Returns the updated agent ticket table partial.
    """
    ticket_ids_str = request.POST.get('ticket_ids', '')
    action = request.POST.get('action', '')
    value = request.POST.get('value', '')
    if not ticket_ids_str or not action:
        return HttpResponse(status=400)

    ids = [int(pk) for pk in ticket_ids_str.split(',') if pk.strip().isdigit()]
    tickets = Ticket.objects.filter(pk__in=ids)

    if action == 'status':
        if value in dict(Ticket.Status.choices):
            for ticket in tickets:
                old_status = ticket.status
                ticket.status = value
                ticket.save()
                TicketActivityLog.objects.create(
                    ticket=ticket, action='status_changed', actor=request.user,
                    details={'from': old_status, 'to': value, 'method': 'bulk'}
                )
    elif action == 'assign':
        if request.user.role not in ['TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
            return HttpResponse(status=403)
        if value.isdigit():
            agent = get_object_or_404(User, pk=int(value))
            for ticket in tickets:
                old_assignee = ticket.assigned_to
                ticket.assigned_to = agent
                ticket.save()
                TicketActivityLog.objects.create(
                    ticket=ticket, action='assigned', actor=request.user,
                    details={'from': old_assignee.get_full_name() if old_assignee else 'Unassigned',
                             'to': agent.get_full_name(), 'method': 'bulk'}
                )

    source = request.POST.get('source', 'unassigned')
    if source == 'assigned':
        tickets = Ticket.objects.filter(
            assigned_to=request.user
        ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED, Ticket.Status.PENDING_APPROVAL]
        ).order_by('-created_at')
    else:
        tickets = Ticket.objects.filter(
            assigned_to__isnull=True
        ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED]
        ).order_by('-created_at')

    assignable_agents = User.objects.filter(role__in=['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']).only('pk', 'first_name', 'last_name', 'email')
    return render(request, 'partials/agent_ticket_table.html', {
        'tickets': tickets,
        'assignable_agents': assignable_agents,
        'status_choices': Ticket.Status.choices,
    })

# ==========================================================================
# TEAM LEAD QUEUE & REASSIGNMENT
# ==========================================================================

@login_required
def team_queue(request):
    """
    Team Lead view: shows all tickets assigned to agents in the same department.
    Allows filtering by individual agent.
    """
    if request.user.role != 'TEAM_LEAD':
        return HttpResponse(status=403)
    team_members = User.objects.filter(department=request.user.department, role='AGENT')
    tickets = Ticket.objects.filter(
        assigned_to__in=team_members
    ).exclude(status__in=[Ticket.Status.RESOLVED, Ticket.Status.CLOSED]
    ).order_by('-created_at')
    agent_id = request.GET.get('agent')
    if agent_id:
        tickets = tickets.filter(assigned_to_id=agent_id)
    context = {
        'tickets': tickets,
        'team_members': team_members,
        'selected_agent': agent_id,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'partials/team_queue.html', context)

@login_required
@require_POST
def team_reassign(request, pk):
    """
    Allows a Team Lead to reassign a ticket to another agent in their team.
    Returns JSON status.
    """
    if request.user.role != 'TEAM_LEAD':
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk)
    new_agent_id = request.POST.get('agent_id')
    agent = get_object_or_404(User, pk=new_agent_id, role='AGENT')
    old_assignee = ticket.assigned_to
    ticket.assigned_to = agent
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket, action='assigned', actor=request.user,
        details={'from': old_assignee.get_full_name() if old_assignee else 'Unassigned',
                 'to': agent.get_full_name()}
    )
    return JsonResponse({'status': 'ok'})

# ==========================================================================
# AUDIT LOG
# ==========================================================================

@login_required
def audit_log(request):
    if request.user.role not in ['TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    logs = TicketActivityLog.objects.select_related('ticket', 'actor').all()
    if request.user.role == 'TEAM_LEAD':
        team_members = User.objects.filter(department=request.user.department, role='AGENT')
        logs = logs.filter(
            Q(ticket__assigned_to__in=team_members) | Q(ticket__requester__in=team_members)
        )
    
    action = request.GET.get('action')
    ticket_id = request.GET.get('ticket')
    if action:
        logs = logs.filter(action=action)
    if ticket_id:
        logs = logs.filter(ticket__number__icontains=ticket_id)
    logs = logs.order_by('-created_at')

    # --- Export logic (CSV, JSON, Excel) ---
    export_format = request.GET.get('export')
    if export_format:
        filename = f"audit_log_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Convert each log to a flat dict for export
        export_data = []
        for log in logs:
            export_data.append({
                'time': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'ticket': log.ticket.number if log.ticket else '—',
                'action': log.action,
                'actor': log.actor.get_full_name() if log.actor else 'System',
                'details': str(log.details) if log.details else ''  # Convert dict to string
            })
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Time', 'Ticket', 'Action', 'Actor', 'Details'])
            for row in export_data:
                writer.writerow([row['time'], row['ticket'], row['action'], row['actor'], row['details']])
            return response

        elif export_format == 'json':
            response = HttpResponse(json.dumps(export_data, indent=2), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
            return response

        elif export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Log"
            ws.append(['Time', 'Ticket', 'Action', 'Actor', 'Details'])
            for row in export_data:
                ws.append([row['time'], row['ticket'], row['action'], row['actor'], row['details']])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response

    # --- Pagination (100 per page) ---
    paginator = Paginator(logs, 50)  # 50 per page (adjust as needed)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'logs': page_obj,
        'action_choices': ['created', 'status_changed', 'assigned', 'unassigned', 'commented', 'remote_session_requested', 'remote_session_status_change'],
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'partials/audit_log.html', context)

# ==========================================================================
# REPORTS DASHBOARD (Charts & KPIs)
# ==========================================================================

@login_required
def reports_dashboard(request):
    """Renders the reports page with SLA compliance, ticket volume, and priority charts."""
    user = request.user
    if user.role not in ['ADMIN', 'SUPERADMIN', 'TEAM_LEAD']:
        return HttpResponse(status=403)
    
    if user.role == 'TEAM_LEAD':
        team_members = User.objects.filter(department=user.department, role='AGENT')
        ticket_filter = Q(assigned_to__in=team_members) | Q(requester__in=team_members)
    else:
        ticket_filter = Q()
    
    # ========== SLA COMPLIANCE ==========
    slas = SLA.objects.all().order_by('priority')
    sla_data = []
    for sla in slas:
        resolved = Ticket.objects.filter(
            ticket_filter, priority=sla.priority,
            status__in=['RESOLVED', 'CLOSED'], resolved_at__isnull=False
        )
        total = resolved.count()
        if total == 0:
            compliance = 100
        else:
            compliant = sum(
                1 for t in resolved
                if t.resolved_at and (t.resolved_at - t.created_at).total_seconds() / 60 <= sla.resolution_minutes
            )
            compliance = round((compliant / total) * 100, 1)
        sla_data.append({'priority': sla.get_priority_display(), 'compliance': compliance})

    # ========== TICKET VOLUME (30 days) ==========
    end = timezone.now().date()
    start = end - timedelta(days=29)  # <-- Use timedelta from datetime module
    
    created_qs = Ticket.objects.filter(
        ticket_filter, created_at__date__gte=start
    ).annotate(date=TruncDate('created_at')).values('date').annotate(count=Count('id')).order_by('date')
    
    resolved_qs = Ticket.objects.filter(
        ticket_filter, resolved_at__isnull=False, resolved_at__date__gte=start
    ).annotate(date=TruncDate('resolved_at')).values('date').annotate(count=Count('id')).order_by('date')

    dates, created_counts, resolved_counts = [], [], []
    for i in range(30):
        d = start + timedelta(days=i)
        dates.append(d.strftime('%m/%d'))
        created_counts.append(next((x['count'] for x in created_qs if x['date'] == d), 0))
        resolved_counts.append(next((x['count'] for x in resolved_qs if x['date'] == d), 0))

    # ========== MTTR ==========
    resolved = Ticket.objects.filter(
        ticket_filter, status__in=['RESOLVED', 'CLOSED'], resolved_at__isnull=False
    )
    mttr = resolved.aggregate(avg_mttr=Avg(F('resolved_at') - F('created_at')))['avg_mttr']
    mttr_minutes = round(mttr.total_seconds() / 60) if mttr else 0

    # ========== BACKLOG ==========
    backlog = Ticket.objects.filter(
        ticket_filter,
        status__in=['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR'],
        created_at__lt=timezone.now() - timedelta(days=7)
    ).count()

    # ========== OPEN BY PRIORITY ==========
    open_by_priority = Ticket.objects.filter(
        ticket_filter,
        status__in=['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']
    ).values('priority').annotate(count=Count('id')).order_by('priority')

    open_labels, open_data = [], []
    for p in open_by_priority:
        open_labels.append(dict(Ticket.Priority.choices)[p['priority']])
        open_data.append(p['count'])

    # ========== ASSET METRICS ==========
    from apps.tickets.models import Asset
    from datetime import date
    
    # Total assets
    total_assets = Asset.objects.count()
    
    # Asset status distribution
    asset_status_labels = ['Active', 'In Store', 'Maintenance', 'Damaged', 'Scrapped']
    asset_status_counts = [
        Asset.objects.filter(status='ACTIVE').count(),
        Asset.objects.filter(status='IN_STORE').count(),
        Asset.objects.filter(status='MAINTENANCE').count(),
        Asset.objects.filter(status='DAMAGED').count(),
        Asset.objects.filter(status='SCRAPPED').count(),
    ]
    
    # Asset fulfillment metrics
    total_asset_requests = Ticket.objects.filter(
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True
    ).count()
    
    fulfilled_asset_requests = Ticket.objects.filter(
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True,
        fulfilled_at__isnull=False
    ).count()
    
    fulfillment_rate = round((fulfilled_asset_requests / total_asset_requests * 100), 1) if total_asset_requests > 0 else 0
    
    # Average fulfillment time (in hours)
    fulfilled_tickets = Ticket.objects.filter(
        type=Ticket.Type.SERVICE_REQUEST,
        is_asset_request=True,
        fulfilled_at__isnull=False,
        created_at__isnull=False
    )
    
    total_hours = 0
    count = 0
    for ticket in fulfilled_tickets:
        delta = ticket.fulfilled_at - ticket.created_at
        total_hours += delta.total_seconds() / 3600
        count += 1
    
    avg_fulfillment_hours = round(total_hours / count, 1) if count > 0 else 0

    context = {
        'sla_data': sla_data,
        'volume_dates': dates,
        'volume_created': created_counts,
        'volume_resolved': resolved_counts,
        'mttr_minutes': mttr_minutes,
        'backlog': backlog,
        'open_priority_labels': open_labels,
        'open_priority_data': open_data,
        'open_total': sum(open_data),
        'sidebar_template': get_sidebar_template(request.user),
        # Asset metrics
        'total_assets': total_assets,
        'asset_status_labels': asset_status_labels,
        'asset_status_counts': asset_status_counts,
        'total_asset_requests': total_asset_requests,
        'fulfilled_asset_requests': fulfilled_asset_requests,
        'fulfillment_rate': fulfillment_rate,
        'avg_fulfillment_hours': avg_fulfillment_hours,
    }
    return render(request, 'dashboards/reports.html', context)

# ==========================================================================
# EXTERNAL CRON TRIGGERS (SLA & CLEANUP)
# ==========================================================================

def is_admin(user): return user.role in ['ADMIN', 'SUPERADMIN']

@login_required
@user_passes_test(is_admin)
def trigger_sla_processing(request):
    """
    Admin-only endpoint to trigger SLA processing.
    Protected by authentication and role check.
    """
    try:
        call_command('process_sla')
        return JsonResponse({'status': 'ok', 'message': 'SLA processing triggered successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@user_passes_test(is_admin)
def trigger_cleanup(request):
    """
    Admin-only endpoint to trigger cleanup of inactive users.
    Protected by authentication and role check.
    """
    try:
        call_command('cleanup_inactive_users')
        return JsonResponse({'status': 'ok', 'message': 'Cleanup triggered successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# If you still need an external trigger (e.g., for cron jobs), use a secure token:
# Option: Use a secure token stored in environment variables
@csrf_exempt
def trigger_sla_processing_external(request):
    """
    External endpoint for cron jobs. Protected by a secure token.
    Token should be set in environment variables, not hardcoded.
    """
    import os
    secret = request.GET.get('secret', '')
    expected_secret = os.environ.get('SLA_TRIGGER_SECRET')
    
    if not expected_secret:
        return JsonResponse({'error': 'SLA trigger not configured'}, status=500)
    
    if secret != expected_secret:
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        call_command('process_sla')
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ==========================================================================
# PLACEHOLDER / STATIC PAGES
# ==========================================================================

@login_required
def catalogue(request):
    if request.user.role not in ['ADMIN', 'SUPERADMIN']: return HttpResponse(status=403)
    return render(request, 'admin/catalogue.html', {'sidebar_template': get_sidebar_template(request.user)})

@login_required
def connectors(request):
    """
    Admin/Superadmin configuration page for remote connectors (Quick Assist, etc.).
    Lists all connectors and allows editing.
    """
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    connectors_list = RemoteConnector.objects.all().order_by('name')
    return render(request, 'admin/connectors.html', {
        'connectors': connectors_list,
        'sidebar_template': get_sidebar_template(request.user),
    })

@login_required
@require_http_methods(['GET', 'POST'])
def connector_edit(request, pk):
    """
    Edit a specific remote connector: enable/disable and update instructions.
    """
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    connector = get_object_or_404(RemoteConnector, pk=pk)
    if request.method == 'POST':
        connector.is_active = request.POST.get('is_active') == 'on'
        connector.instructions_for_requester = request.POST.get('instructions_for_requester', '')
        connector.instructions_for_agent = request.POST.get('instructions_for_agent', '')
        connector.save()
        return redirect('tickets:connectors')
    return render(request, 'admin/connector_form.html', {
        'connector': connector,
        'sidebar_template': get_sidebar_template(request.user),
    })

# ==========================================================================
# ASSET MANAGEMENT
# ==========================================================================

# ==========================================================================
# HELPER: Parse date for asset import
# ==========================================================================

def parse_date(value):
    """Parse date from various formats for asset import."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%b %d, %Y']:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None

# ==========================================================================
# ASSET LIST
# ==========================================================================

@login_required
def assets(request):
    if request.user.role not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD']:
        return HttpResponse(status=403)

    # Use renamed filter parameters
    query = request.GET.get('filter_q', '')
    asset_type = request.GET.get('filter_type', '')
    status_filter = request.GET.get('filter_status', '')
    location_filter = request.GET.get('filter_location', '')

    assets_list = Asset.objects.all().order_by('-created_at')

    if query:
        assets_list = assets_list.filter(
            Q(name__icontains=query) |
            Q(tracking_id__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(model__icontains=query) |
            Q(manufacturer__icontains=query)
        )
    if asset_type:
        assets_list = assets_list.filter(asset_type=asset_type)
    if status_filter:
        assets_list = assets_list.filter(status=status_filter)
    if location_filter:
        assets_list = assets_list.filter(location__icontains=location_filter)

    paginator = Paginator(assets_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    users = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    context = {
        'assets': page_obj,
        'users': users,
        'asset_types': Asset.AssetType.choices,
        'asset_type_values': [v for v, _ in Asset.AssetType.choices],
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'location_choices': Asset.Location.choices,
        'location_values': [v for v, _ in Asset.Location.choices],
        'query': query,
        'selected_type': asset_type,
        'selected_status': status_filter,
        'selected_location': location_filter,
        'sidebar_template': get_sidebar_template(request.user),
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/asset_table.html', context)

    return render(request, 'tickets/asset_list.html', context)


# ==========================================================================
# ASSET CREATE PAGE (Dedicated Page)
# ==========================================================================

@login_required
@require_http_methods(['GET', 'POST'])
def asset_create_page(request):
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.tracking_id = None
            asset.save()
            
            # Create CREATED log
            AssetLog.objects.create(
                asset=asset,
                action=AssetLog.Action.CREATED,
                actor=request.user,
                details={'name': asset.name, 'type': asset.asset_type}
            )
            
            # Create ASSIGNED log if assigned to someone
            if asset.assigned_to:
                AssetLog.objects.create(
                    asset=asset,
                    action=AssetLog.Action.ASSIGNED,
                    actor=request.user,
                    details={
                        'from': None,
                        'to': asset.assigned_to.get_full_name() if asset.assigned_to else None,
                        'comment': 'Initial assignment'
                    }
                )
            
            messages.success(request, f'Asset "{asset.name}" created successfully!')
            return redirect('tickets:assets')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssetForm()

    context = {
        'form': form,
        'asset': None,
        'action': 'create',
        'users': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'asset_types': Asset.AssetType.choices,
        'asset_type_values': [v for v, _ in Asset.AssetType.choices],
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'location_choices': Asset.Location.choices,
        'location_values': [v for v, _ in Asset.Location.choices],
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/asset_form_page.html', context)


# ==========================================================================
# ASSET EDIT PAGE (Dedicated Page)
# ==========================================================================

@login_required
@require_http_methods(['GET', 'POST'])
def asset_edit_page(request, pk):
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)

    if request.method == 'POST':
        # Store the old assigned_to value before saving
        old_assigned_to = asset.assigned_to
        
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            asset = form.save()
            
            # Check if assigned_to changed
            new_assigned_to = asset.assigned_to
            
            # Create appropriate logs based on what changed
            if old_assigned_to != new_assigned_to:
                # This is a reassignment or unassignment
                if new_assigned_to:
                    # Reassigned to someone
                    AssetLog.objects.create(
                        asset=asset,
                        action=AssetLog.Action.ASSIGNED,
                        actor=request.user,
                        details={
                            'from': old_assigned_to.get_full_name() if old_assigned_to else None,
                            'to': new_assigned_to.get_full_name() if new_assigned_to else None,
                            'comment': 'Reassigned via edit form'
                        }
                    )
                else:
                    # Unassigned
                    AssetLog.objects.create(
                        asset=asset,
                        action=AssetLog.Action.UNASSIGNED,
                        actor=request.user,
                        details={
                            'from': old_assigned_to.get_full_name() if old_assigned_to else None,
                            'to': None,
                            'comment': 'Unassigned via edit form'
                        }
                    )
            else:
                # General update (no assignment change)
                AssetLog.objects.create(
                    asset=asset,
                    action=AssetLog.Action.UPDATED,
                    actor=request.user,
                    details={'source': 'edit_page'}
                )
            
            messages.success(request, f'Asset "{asset.name}" updated successfully!')
            
            # Preserve filters when redirecting back
            source = request.GET.get('source', 'list')
            query = request.GET.get('q', '')
            asset_type = request.GET.get('type', '')
            status = request.GET.get('status', '')
            location = request.GET.get('location', '')
            
            redirect_url = reverse('tickets:assets')
            if source == 'list':
                params = []
                if query:
                    params.append(f'q={query}')
                if asset_type:
                    params.append(f'type={asset_type}')
                if status:
                    params.append(f'status={status}')
                if location:
                    params.append(f'location={location}')
                if params:
                    redirect_url += '?' + '&'.join(params)
            
            return redirect(redirect_url)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssetForm(instance=asset)

    context = {
        'form': form,
        'asset': asset,
        'action': 'edit',
        'users': User.objects.filter(is_active=True).order_by('first_name', 'last_name'),
        'asset_types': Asset.AssetType.choices,
        'asset_type_values': [v for v, _ in Asset.AssetType.choices],
        'status_choices': Asset.Status.choices,
        'status_values': [v for v, _ in Asset.Status.choices],
        'location_choices': Asset.Location.choices,
        'location_values': [v for v, _ in Asset.Location.choices],
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/asset_form_page.html', context)

# ==========================================================================
# ASSET REASSIGN
# ==========================================================================

@login_required
@require_POST
def asset_reassign(request, pk):
    if request.user.role not in ['ADMIN', 'SUPERADMIN', 'TEAM_LEAD']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    new_user_id = request.POST.get('assigned_to')
    comment = request.POST.get('comment', '')

    old_user = asset.assigned_to
    if new_user_id:
        new_user = get_object_or_404(User, pk=new_user_id)
        asset.assigned_to = new_user
    else:
        asset.assigned_to = None
    asset.save()

    AssetLog.objects.create(
        asset=asset,
        action=AssetLog.Action.ASSIGNED if new_user_id else AssetLog.Action.UNASSIGNED,
        actor=request.user,
        details={
            'from': old_user.get_full_name() if old_user else None,
            'to': new_user.get_full_name() if new_user_id else None,
            'comment': comment
        }
    )

    return redirect('tickets:assets')

# ==========================================================================
# ASSET DETAIL
# ==========================================================================

@login_required
def asset_detail(request, pk):
    if request.user.role not in ['ADMIN', 'SUPERADMIN', 'AGENT', 'TEAM_LEAD']:
        return HttpResponse(status=403)
    
    asset = get_object_or_404(Asset, pk=pk)
    logs = asset.logs.all()[:10]  # Recent activity
    
    return render(request, 'tickets/asset_detail.html', {
        'asset': asset,
        'logs': logs,
        'sidebar_template': get_sidebar_template(request.user),
    })

# ==========================================================================
# ASSET SCRAP REQUEST
# ==========================================================================

@login_required
@require_POST
def asset_scrap_request(request, pk):
    if request.user.role not in ['ADMIN', 'SUPERADMIN', 'TEAM_LEAD']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    comment = request.POST.get('comment', '')

    if asset.status == Asset.Status.SCRAPPED:
        return JsonResponse({'error': 'Asset already scrapped.'}, status=400)

    asset.status = Asset.Status.DAMAGED
    asset.save()

    AssetLog.objects.create(
        asset=asset,
        action=AssetLog.Action.SCRAP_REQUESTED,
        actor=request.user,
        details={'comment': comment}
    )

    return redirect('tickets:assets')

# ==========================================================================
# ASSET SCRAP APPROVE
# ==========================================================================

@login_required
@require_POST
def asset_scrap_approve(request, pk):
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    asset = get_object_or_404(Asset, pk=pk)
    action = request.POST.get('action')  # 'approve' or 'reject'

    if action == 'approve':
        asset.status = Asset.Status.SCRAPPED
        asset.scrap_approved = True
        asset.scrap_approved_at = timezone.now()
        asset.scrap_approved_by = request.user
        asset.save()
        AssetLog.objects.create(
            asset=asset,
            action=AssetLog.Action.SCRAP_APPROVED,
            actor=request.user,
            details={'comment': request.POST.get('comment', '')}
        )
    else:
        # Reject: revert to previous status (e.g., ACTIVE or IN_STORE)
        asset.status = Asset.Status.ACTIVE  # or a configurable fallback
        asset.save()
        AssetLog.objects.create(
            asset=asset,
            action=AssetLog.Action.SCRAP_REJECTED,
            actor=request.user,
            details={'comment': request.POST.get('comment', '')}
        )

    return redirect('tickets:assets')

# ==========================================================================
# ASSET CALCULATE WARRANTY
# ==========================================================================

@login_required
def asset_calculate_warranty(request):
    """Calculate warranty expiry date based on purchase date and duration."""
    purchase_date_str = request.GET.get('purchase_date')
    duration_years = request.GET.get('warranty_duration', 0)
    expiry_date_str = ''
    
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Warranty calculation: purchase_date={purchase_date_str}, duration={duration_years}")
    
    try:
        duration_years = int(duration_years)
        if duration_years > 0 and purchase_date_str:
            from datetime import datetime
            purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
            # Calculate expiry by adding years
            try:
                expiry_date = purchase_date.replace(year=purchase_date.year + duration_years)
                expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                logger.info(f"Warranty calculation result: {expiry_date_str}")
            except ValueError:
                # Handle Feb 29 edge case - approximate by adding days
                from datetime import timedelta
                days = 365 * duration_years
                expiry_date = purchase_date + timedelta(days=days)
                expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                logger.info(f"Warranty calculation (approx): {expiry_date_str}")
    except (ValueError, TypeError) as e:
        logger.error(f"Warranty calculation error: {e}")
        pass
    
    return render(request, 'partials/warranty_expiry_input.html', {'value': expiry_date_str})

# ==========================================================================
# REMOTE SESSION REQUESTS (Quick Assist integration)
# ==========================================================================

@login_required
@require_POST
def request_remote_session(request, pk):
    """
    Initiates a remote session request from an agent to the ticket requester.
    - Creates a RemoteSession record (status=REQUESTED).
    - Adds a public comment on the ticket.
    - Sends an in‑app notification to the requester.
    - Sends an email to the requester with a link to accept the session.
    - Logs the action in the activity log.
    Only agents, team leads, admins, and superadmins can call this.
    """
    ticket = get_object_or_404(Ticket, pk=pk)
    if request.user.role not in [User.Role.AGENT, User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)

    # Get the first active remote connector (e.g., Quick Assist)
    connector = RemoteConnector.objects.filter(is_active=True).first()
    if not connector:
        messages.error(request, 'No active remote connector configured. Please contact your administrator.')
        return redirect('tickets:conversation', pk=pk)

    # Check if there's already a pending or active session for this ticket
    existing = RemoteSession.objects.filter(ticket=ticket, status__in=['REQUESTED', 'ACCEPTED', 'STARTED']).first()
    if existing:
        return JsonResponse({'error': 'A remote session is already pending or in progress.'}, status=400)
    
    session = RemoteSession.objects.create(
        ticket=ticket,
        requester=ticket.requester,
        agent=request.user,
        connector=connector,
        status='REQUESTED'
    )

    # Add a public comment to the ticket timeline
    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"Remote session requested via {connector.name}. Please check your notifications to accept.",
        visibility='PUBLIC'
    )

    # Send in‑app notification
    Notification.objects.create(
        recipient=ticket.requester,
        message=f"Remote session requested for ticket {ticket.number}. Click to accept.",
        url=reverse('tickets:remote_session_detail', args=[session.pk]),
        type=Notification.Type.REMOTE_SESSION
    )

    # Send email notification to the requester
    accept_url = request.build_absolute_uri(reverse('tickets:remote_session_detail', args=[session.pk]))
    reject_url = request.build_absolute_uri(reverse('tickets:remote_session_detail', args=[session.pk])) + '?action=reject'

    html_message = render_to_string('emails/remote_session_request.html', {
        'requester_name': ticket.requester.get_full_name() or ticket.requester.email,
        'ticket_number': ticket.number,
        'accept_url': accept_url,
        'reject_url': reject_url,
    })
    plain_message = strip_tags(html_message)

    send_mail(
        subject=f"Remote Session Request – Ticket {ticket.number}",
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[ticket.requester.email],
        html_message=html_message,
        fail_silently=True,
    )
    
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='remote_session_requested',
        actor=request.user,
        details={'connector': connector.name, 'session_id': session.pk}
    )
    
    return JsonResponse({'session_id': session.pk, 'status': 'requested'})

@login_required
def remote_session_detail(request, session_pk):
    session = get_object_or_404(RemoteSession, pk=session_pk)
    user = request.user
    if user != session.requester and user != session.agent:
        return HttpResponse(status=403)
    
    if user == session.agent:
        instructions = session.connector.instructions_for_agent
        role = 'agent'
    else:
        instructions = session.connector.instructions_for_requester
        role = 'requester'
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(RemoteSession.STATUS_CHOICES):
            old_status = session.status
            # Handle REJECT from requester
            if new_status == 'REJECTED' and role == 'requester':
                session.status = 'REJECTED'
                session.save()
                # Notify agent
                Notification.objects.create(
                    recipient=session.agent,
                    message=f"{session.requester.get_full_name()} rejected the remote session for ticket {session.ticket.number}.",
                    url=reverse('tickets:remote_session_detail', args=[session.pk])
                )
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': 'REJECTED', 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle ACCEPT from requester
            elif new_status == 'ACCEPTED' and role == 'requester' and old_status == 'REQUESTED':
                session.status = 'ACCEPTED'
                session.save()
                # Notify agent
                Notification.objects.create(
                    recipient=session.agent,
                    message=f"{session.requester.get_full_name()} accepted the remote session for ticket {session.ticket.number}.",
                    url=reverse('tickets:remote_session_detail', args=[session.pk]),
                    type=Notification.Type.REMOTE_SESSION
                )
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': 'ACCEPTED', 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle START with code from agent
            elif new_status == 'STARTED' and role == 'agent' and old_status == 'ACCEPTED':
                code = request.POST.get('quick_assist_code', '').strip()
                if not code or len(code) < 6:
                    # Optionally show error message
                    pass
                else:
                    session.session_code = code
                    session.status = 'STARTED'
                    session.started_at = timezone.now()
                    session.save()
                    # Send automatic public comment on ticket
                    TicketComment.objects.create(
                        ticket=session.ticket,
                        author=request.user,
                        body=f"Remote session code: {code}. Please use this code in Quick Assist to start the session.",
                        visibility='PUBLIC'
                    )
                    # Send email to requester
                    html_message = render_to_string('emails/remote_session_code.html', {
                        'requester_name': session.requester.get_full_name() or session.requester.email,
                        'ticket_number': session.ticket.number,
                        'code': code,
                    })
                    plain_message = strip_tags(html_message)

                    send_mail(
                        subject=f"Remote Session Code – Ticket {session.ticket.number}",
                        message=plain_message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[session.requester.email],
                        html_message=html_message,
                        fail_silently=True,
                    )
                    TicketActivityLog.objects.create(
                        ticket=session.ticket,
                        action='remote_session_status_change',
                        actor=user,
                        details={'from': old_status, 'to': 'STARTED', 'session_id': session.pk, 'code': code}
                    )
                    return redirect('tickets:remote_session_detail', session_pk=session.pk)
            
            # Handle END from agent
            elif new_status == 'ENDED' and role == 'agent' and old_status == 'STARTED':
                session.status = 'ENDED'
                session.ended_at = timezone.now()
                session.save()
                TicketActivityLog.objects.create(
                    ticket=session.ticket,
                    action='remote_session_status_change',
                    actor=user,
                    details={'from': old_status, 'to': 'ENDED', 'session_id': session.pk}
                )
                return redirect('tickets:remote_session_detail', session_pk=session.pk)
    
    context = {
        'session': session,
        'instructions': instructions,
        'role': role,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/remote_session_detail.html', context)

@login_required
def remote_session_pending_count(request):
    """
    Returns the count of pending remote sessions for the current user.
    For agents: sessions they requested with status REQUESTED or ACCEPTED.
    For requesters: sessions requested for their tickets with status REQUESTED.
    """
    user = request.user
    if user.role in ['ADMIN', 'SUPERADMIN']:
        # Admins see all pending sessions
        count = RemoteSession.objects.filter(status__in=['REQUESTED', 'ACCEPTED']).count()
    elif user.role in ['AGENT', 'TEAM_LEAD']:
        count = RemoteSession.objects.filter(agent=user, status__in=['REQUESTED', 'ACCEPTED']).count()
    else:
        # End users: sessions requested for their tickets
        count = RemoteSession.objects.filter(requester=user, status='REQUESTED').count()
    return render(request, 'partials/remote_session_badge.html', {'count': count})

@login_required
def remote_sessions_list(request):
    """
    List all remote sessions relevant to the logged‑in user.
    - Agents see sessions they initiated (as agent).
    - Requesters see sessions requested for their tickets.
    - Admins/Superadmins see all sessions (optional, but we'll filter by role).
    """
    user = request.user
    if user.role in ['ADMIN', 'SUPERADMIN']:
        sessions = RemoteSession.objects.all().order_by('-created_at')
    elif user.role in ['AGENT', 'TEAM_LEAD']:
        sessions = RemoteSession.objects.filter(agent=user).order_by('-created_at')
    else:
        sessions = RemoteSession.objects.filter(requester=user).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(sessions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'sessions': page_obj,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'tickets/remote_sessions_list.html', context)

@login_required
def escalated_tickets(request):
    if request.user.role not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)

    tickets = Ticket.objects.filter(status=Ticket.Status.ESCALATED).order_by('-created_at')

    # If Team Lead, filter by their department
    if request.user.role == User.Role.TEAM_LEAD:
        tickets = tickets.filter(assigned_to__department=request.user.department)

    # Agents in the same department (for reassign)
    agents = User.objects.filter(department=request.user.department, role=User.Role.AGENT, is_active=True)

    # Workload: open tickets per agent
    open_statuses = ['NEW', 'TRIAGED', 'ASSIGNED', 'IN_PROGRESS', 'PENDING_USER', 'PENDING_VENDOR']
    agent_workload = {}
    for agent in agents:
        agent_workload[agent.pk] = Ticket.objects.filter(
            assigned_to=agent,
            status__in=open_statuses
        ).count()

    # Reason macros
    reassign_reasons = Macro.objects.filter(type=Macro.Type.REASSIGN_REASON)
    return_reasons = Macro.objects.filter(type=Macro.Type.RETURN_REASON)

    context = {
        'tickets': tickets,
        'assignable_agents': agents,
        'agent_workload': agent_workload,
        'reassign_reasons': reassign_reasons,
        'return_reasons': return_reasons,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'team_lead/escalated_tickets.html', context)

@login_required
@require_POST
def reassign_escalated(request, pk):
    if request.user.role not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.ESCALATED)
    agent_id = request.POST.get('agent_id')
    comment = request.POST.get('comment', '')
    agent = get_object_or_404(User, pk=agent_id, role=User.Role.AGENT)
    ticket.assigned_to = agent
    ticket.status = Ticket.Status.ASSIGNED
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='reassigned_escalated',
        actor=request.user,
        details={'to': agent.get_full_name(), 'comment': comment}
    )
    # Notify agent
    Notification.objects.create(
        recipient=agent,
        message=f"Ticket {ticket.number} has been reassigned to you by {request.user.get_full_name()}.",
        url=reverse('tickets:detail', args=[ticket.pk])
    )
    return redirect('tickets:escalated_tickets')

@login_required
@require_POST
def return_escalated_to_pool(request, pk):
    if request.user.role not in [User.Role.TEAM_LEAD, User.Role.ADMIN, User.Role.SUPERADMIN]:
        return HttpResponse(status=403)
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.ESCALATED)
    comment = request.POST.get('comment', '')
    if not comment:
        return JsonResponse({'error': 'Comment is required'}, status=400)
    ticket.assigned_to = None
    ticket.status = Ticket.Status.NEW
    ticket.save()
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='returned_to_pool',
        actor=request.user,
        details={'comment': comment}
    )
    # Notify? optional
    return redirect('tickets:escalated_tickets')


def kb_suggestions(request):
    return render(request, 'partials/kb_suggestions.html', {'articles': []})

# ==========================================================================
# ATTACHMENT PREVIEW AND DOWNLOAD
# ==========================================================================

@login_required
def attachment_preview(request, pk):
    """
    Returns a modal with a preview of the attachment.
    Supports images, PDF, Office documents, and text files.
    """
    attachment = get_object_or_404(Attachment, pk=pk)
    ticket = attachment.ticket

    # Permission check: same as download
    if request.user != ticket.requester and request.user.role not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)

    # Get file URL for embedding
    file_url = request.build_absolute_uri(attachment.file.url)
    content_type = attachment.content_type or ''
    filename = attachment.filename

    # Determine preview type
    preview_type = 'unknown'
    embed_url = None
    text_content = None

    if content_type.startswith('image/'):
        preview_type = 'image'
    elif content_type == 'application/pdf':
        preview_type = 'pdf'
        embed_url = f"https://docs.google.com/gview?url={file_url}&embedded=true"
    elif content_type in [
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-powerpoint',
        'application/vnd.openxmlformats-officedocument.presentationml.presentation'
    ]:
        preview_type = 'office'
        embed_url = f"https://docs.google.com/gview?url={file_url}&embedded=true"
    elif content_type.startswith('text/') or filename.endswith(('.txt', '.csv', '.log', '.py', '.js', '.html', '.css')):
        preview_type = 'text'
        # Fetch the file content (for small text files only)
        try:
            with attachment.file.open('r') as f:
                text_content = f.read()
                # Limit size to prevent huge files
                if len(text_content) > 100000:  # 100KB limit
                    text_content = "File too large to preview as text."
        except Exception:
            text_content = "Could not read file content."

    context = {
        'attachment': attachment,
        'preview_type': preview_type,
        'embed_url': embed_url,
        'text_content': text_content,
        'file_url': file_url,
    }
    return render(request, 'tickets/attachment_preview.html', context)

@login_required
def attachment_download(request, pk):
    """
    Serves a file attachment. Only the ticket requester or agents/leads/admins can download.
    """
    attachment = get_object_or_404(Attachment, pk=pk)
    ticket = attachment.ticket
    if request.user != ticket.requester and request.user.role not in ['AGENT', 'TEAM_LEAD', 'ADMIN', 'SUPERADMIN']:
        raise PermissionDenied
    response = FileResponse(attachment.file.open('rb'), content_type=attachment.content_type or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
    return response

# ==========================================================================
# SLA MANAGEMENT (Admin & Superadmin only)
# ==========================================================================

def is_admin(user): return user.role in ['ADMIN', 'SUPERADMIN']

@login_required
@user_passes_test(is_admin)
def sla_list(request):
    slas = SLA.objects.all().order_by('priority')
    calendars = BusinessCalendar.objects.all()
    rules = EscalationRule.objects.all().order_by('priority', 'timer_type', 'threshold_percent')
    context = {
        'slas': slas,
        'calendars': calendars,
        'rules': rules,
        'priority_choices': Ticket.Priority.choices,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'admin/sla_management.html', context)

@login_required
@user_passes_test(is_admin)
@require_POST
def sla_create(request):
    priority = request.POST.get('priority')
    response = request.POST.get('response_minutes')
    resolution = request.POST.get('resolution_minutes')
    calendar_id = request.POST.get('calendar_id') or None
    SLA.objects.update_or_create(
        priority=priority,
        defaults={'response_minutes': response, 'resolution_minutes': resolution, 'calendar_id': calendar_id}
    )
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def sla_delete(request, pk):
    sla = get_object_or_404(SLA, pk=pk)
    sla.delete()
    return redirect('tickets:sla_management')

@login_required
def sla_badge(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    return render(request, 'partials/sla_badge.html', {'ticket': ticket})

@login_required
@user_passes_test(is_admin)
@require_POST
def calendar_create(request):
    name = request.POST.get('name')
    workdays = request.POST.getlist('workdays')
    work_start = request.POST.get('work_start')
    work_end = request.POST.get('work_end')
    holidays_str = request.POST.get('holidays', '')
    holidays = [h.strip() for h in holidays_str.split(',') if h.strip()]
    BusinessCalendar.objects.create(
        name=name, workdays=workdays, work_start=work_start, work_end=work_end, holidays=holidays
    )
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def rule_create(request):
    priority = request.POST.get('priority')
    timer_type = request.POST.get('timer_type')
    threshold = request.POST.get('threshold_percent')
    action = request.POST.get('action_type')
    notify_role = request.POST.get('notify_role') if action == 'notify' else None
    reassign_to_role = request.POST.get('reassign_to_role') if action == 'reassign' else None
    EscalationRule.objects.create(
        priority=priority, timer_type=timer_type, threshold_percent=threshold,
        action_type=action, notify_role=notify_role, reassign_to_role=reassign_to_role
    )
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def rule_delete(request, pk):
    rule = get_object_or_404(EscalationRule, pk=pk)
    rule.delete()
    return redirect('tickets:sla_management')

@login_required
@user_passes_test(is_admin)
@require_POST
def calendar_delete(request, pk):
    cal = get_object_or_404(BusinessCalendar, pk=pk)
    cal.delete()
    return redirect('tickets:sla_management')

# ==========================================================================
# MANAGER WORKFLOW (Team Lead reviews service requests)
# ==========================================================================

@login_required
def manager_review_queue(request):
    """Team Lead view – list service requests pending manager review."""
    if request.user.role != User.Role.TEAM_LEAD:
        return HttpResponse(status=403)

    tickets = Ticket.objects.filter(
        status=Ticket.Status.PENDING_MANAGER_REVIEW,
        requester__department=request.user.department
    ).order_by('-created_at')

    context = {
        'tickets': tickets,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'team_lead/manager_review_queue.html', context)


@login_required
@login_required
def manager_review_ticket(request, pk):
    """Team Lead review page for a single service request."""
    if request.user.role != User.Role.TEAM_LEAD:
        return HttpResponse(status=403)

    ticket = get_object_or_404(Ticket, pk=pk)

    # Security: ensure ticket belongs to Team Lead's department
    if ticket.requester.department != request.user.department:
        return HttpResponse(status=403)

    # Only allow review of PENDING_MANAGER_REVIEW tickets
    if ticket.status != Ticket.Status.PENDING_MANAGER_REVIEW:
        messages.warning(request, f'Ticket {ticket.number} is not pending manager review.')
        return redirect('tickets:manager_review_queue')

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        comment = request.POST.get('comment', '').strip()

        if not comment:
            messages.error(request, 'Please provide a comment.')
            return redirect('tickets:manager_review_ticket', pk=pk)

        if action == 'approve':
            # ================================================================
            # ASSET REQUEST ROUTING
            # ================================================================
            if ticket.is_asset_request:
                # Asset requests go to PENDING_FULFILLMENT for Admin to fulfill
                ticket.status = Ticket.Status.PENDING_FULFILLMENT
                ticket.save()
                
                TicketActivityLog.objects.create(
                    ticket=ticket,
                    action='manager_approved',
                    actor=request.user,
                    details={'comment': comment, 'routed_to': 'PENDING_FULFILLMENT'}
                )
                
                # Notify Admins about pending fulfillment
                admins = User.objects.filter(role=User.Role.ADMIN, is_active=True)
                for admin in admins:
                    Notification.objects.create(
                        recipient=admin,
                        message=f'Asset request {ticket.number} from {ticket.requester.get_full_name()} needs fulfillment.',
                        url=reverse('tickets:conversation', args=[ticket.pk])
                    )
                
                # Notify requester
                Notification.objects.create(
                    recipient=ticket.requester,
                    message=f'Your asset request {ticket.number} has been approved by your manager and is pending fulfillment.',
                    url=reverse('tickets:detail', args=[ticket.pk])
                )
                
                messages.success(request, f'Asset request {ticket.number} approved. An admin will fulfill it shortly.')
            
            else:
                # ================================================================
                # NON-ASSET REQUESTS: Go directly to APPROVED
                # Skip the Approver role entirely
                # ================================================================
                ticket.status = Ticket.Status.APPROVED
                ticket.save()
                
                TicketActivityLog.objects.create(
                    ticket=ticket,
                    action='manager_approved',
                    actor=request.user,
                    details={'comment': comment, 'routed_to': 'APPROVED'}
                )
                
                # Notify requester
                Notification.objects.create(
                    recipient=ticket.requester,
                    message=f'Your service request {ticket.number} has been approved.',
                    url=reverse('tickets:detail', args=[ticket.pk])
                )
                
                # ================================================================
                # Ticket is now APPROVED - send to Agent Pool
                # ================================================================
                # Notify agents about new approved ticket
                agents = User.objects.filter(role__in=[User.Role.AGENT, User.Role.TEAM_LEAD])
                for agent in agents:
                    Notification.objects.create(
                        recipient=agent,
                        message=f'New approved ticket {ticket.number}: {ticket.title}',
                        url=reverse('tickets:detail', args=[ticket.pk])
                    )
                
                messages.success(request, f'Ticket {ticket.number} approved and sent to agent queue.')

        elif action == 'reject':
            ticket.status = Ticket.Status.CLOSED
            ticket.save()
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='manager_rejected',
                actor=request.user,
                details={'comment': comment}
            )
            Notification.objects.create(
                recipient=ticket.requester,
                message=f'Your service request {ticket.number} was rejected by your manager. Reason: {comment}',
                url=reverse('tickets:detail', args=[ticket.pk])
            )
            messages.info(request, f'Ticket {ticket.number} rejected.')

        elif action == 'request_changes':
            ticket.status = Ticket.Status.PENDING_USER
            ticket.save()
            TicketActivityLog.objects.create(
                ticket=ticket,
                action='manager_requested_changes',
                actor=request.user,
                details={'comment': comment}
            )
            Notification.objects.create(
                recipient=ticket.requester,
                message=f'Changes requested for ticket {ticket.number} by your manager: {comment}',
                url=reverse('tickets:detail', args=[ticket.pk])
            )
            messages.info(request, f'Changes requested on ticket {ticket.number}.')

        else:
            messages.error(request, f'Invalid action: "{action}"')
            return redirect('tickets:manager_review_ticket', pk=pk)

        return redirect('tickets:manager_review_queue')

    # GET – render review page
    comments = ticket.comments.all().order_by('created_at')
    initial_attachments = ticket.attachments.filter(comment__isnull=True)

    context = {
        'ticket': ticket,
        'comments': comments,
        'initial_attachments': initial_attachments,
        'sidebar_template': get_sidebar_template(request.user),
    }
    return render(request, 'team_lead/manager_review_ticket.html', context)

@login_required
def manager_review_count(request):
    if request.user.role != User.Role.TEAM_LEAD:
        return HttpResponse('')
    count = Ticket.objects.filter(
        status=Ticket.Status.PENDING_MANAGER_REVIEW,
        requester__department=request.user.department
    ).count()
    return render(request, 'partials/manager_review_badge.html', {'count': count})

# ==========================================================================
# ASSET EXPORT
# ==========================================================================

@login_required
def asset_export(request):
    """Export assets as CSV, Excel, or JSON"""
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    export_format = request.GET.get('format', 'csv')
    filename = f"assets_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
    
    # Get filtered assets (respect current filters)
    query = request.GET.get('q', '')
    asset_type = request.GET.get('type', '')
    status_filter = request.GET.get('status', '')
    location_filter = request.GET.get('location', '')
    
    assets = Asset.objects.all().order_by('tracking_id')
    
    if query:
        assets = assets.filter(
            Q(name__icontains=query) |
            Q(tracking_id__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(model__icontains=query) |
            Q(manufacturer__icontains=query)
        )
    if asset_type:
        assets = assets.filter(asset_type=asset_type)
    if status_filter:
        assets = assets.filter(status=status_filter)
    if location_filter:
        assets = assets.filter(location__icontains=location_filter)
    
    # Prepare data
    data = []
    for asset in assets:
        data.append({
            'Tracking ID': asset.tracking_id,
            'Name': asset.name,
            'Type': asset.get_asset_type_display(),
            'Serial Number': asset.serial_number,
            'Model': asset.model,
            'Manufacturer': asset.manufacturer,
            'Location': asset.location,
            'Status': asset.get_status_display(),
            'Assigned To': asset.assigned_to.get_full_name() if asset.assigned_to else '',
            'Assigned Department': asset.assigned_to.department if asset.assigned_to else '',
            'Purchase Date': asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
            'Warranty Expiry': asset.warranty_expiry.strftime('%Y-%m-%d') if asset.warranty_expiry else '',
            'Warranty Duration (Years)': asset.warranty_duration_years,
            'Notes': asset.notes,
            'Created': asset.created_at.strftime('%Y-%m-%d %H:%M'),
            'Updated': asset.updated_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    # Export based on format
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.DictWriter(response, fieldnames=data[0].keys() if data else [])
        writer.writeheader()
        writer.writerows(data)
        return response
    
    elif export_format == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        
        if data:
            # Headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
        return response
    
    elif export_format == 'json':
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        return response
    
    return HttpResponse('Invalid format', status=400)


# ==========================================================================
# ASSET IMPORT
# ==========================================================================

@login_required
@require_POST
def asset_import(request):
    """Import assets from CSV or Excel file"""
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Please select a file to import.')
        return redirect('tickets:assets')
    
    # Check file type
    file_name = file.name.lower()
    is_csv = file_name.endswith('.csv')
    is_excel = file_name.endswith(('.xlsx', '.xls'))
    
    if not (is_csv or is_excel):
        messages.error(request, 'Please upload a CSV or Excel file.')
        return redirect('tickets:assets')
    
    imported = 0
    errors = []
    warnings = []
    
    try:
        if is_csv:
            # Parse CSV
            decoded = file.read().decode('utf-8')
            reader = csv.DictReader(decoded.splitlines())
            rows = list(reader)
        else:
            # Parse Excel
            wb = Workbook()
            ws = wb.active
            # Read Excel manually since we're using openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                rows.append(row_dict)
        
        # Import each row
        for row in rows:
            try:
                # Skip empty rows
                if not row.get('Name') and not row.get('name'):
                    continue
                
                # Map columns (case-insensitive)
                name = row.get('Name') or row.get('name')
                asset_type = row.get('Type') or row.get('type')
                serial_number = row.get('Serial Number') or row.get('serial_number') or row.get('Serial') or ''
                model = row.get('Model') or row.get('model') or ''
                manufacturer = row.get('Manufacturer') or row.get('manufacturer') or ''
                location = row.get('Location') or row.get('location') or ''
                status = row.get('Status') or row.get('status') or 'ACTIVE'
                purchase_date = parse_date(row.get('Purchase Date') or row.get('purchase_date'))
                warranty_expiry = parse_date(row.get('Warranty Expiry') or row.get('warranty_expiry'))
                warranty_duration = row.get('Warranty Duration (Years)') or row.get('warranty_duration') or 0
                notes = row.get('Notes') or row.get('notes') or ''
                assigned_to_name = row.get('Assigned To') or row.get('assigned_to') or ''
                
                # Find assigned user by name or email
                assigned_to = None
                if assigned_to_name:
                    assigned_to = User.objects.filter(
                        Q(first_name__icontains=assigned_to_name) |
                        Q(last_name__icontains=assigned_to_name) |
                        Q(email__icontains=assigned_to_name)
                    ).first()
                
                # Create asset
                asset = Asset.objects.create(
                    name=name,
                    asset_type=asset_type or 'OTHER',
                    serial_number=serial_number,
                    model=model,
                    manufacturer=manufacturer,
                    location=location,
                    status=status,
                    purchase_date=purchase_date if purchase_date else None,
                    warranty_expiry=warranty_expiry if warranty_expiry else None,
                    warranty_duration_years=int(warranty_duration) if warranty_duration else 0,
                    notes=notes,
                    assigned_to=assigned_to,
                )
                
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {len(rows)}: {str(e)}")
        
    except Exception as e:
        messages.error(request, f'Error reading file: {str(e)}')
        return redirect('tickets:assets')
    
    # Show results
    if imported > 0:
        messages.success(request, f'✅ Successfully imported {imported} asset(s).')
    if errors:
        messages.warning(request, f'⚠️ {len(errors)} error(s) occurred.')
    if not imported and not errors:
        messages.warning(request, 'No assets were imported. Please check your file format.')
    
    return redirect('tickets:assets')


# ==========================================================================
# ASSET FULFILLMENT (Admin only)
# ==========================================================================

@login_required
def fulfill_asset_modal(request, pk):
    """Returns the fulfillment modal for an asset request."""
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.PENDING_FULFILLMENT)
    
    # Get available assets (IN_STORE or unassigned ACTIVE)
    available_assets = Asset.objects.filter(
        status__in=['IN_STORE', 'ACTIVE'],
        assigned_to__isnull=True
    ).order_by('name')
    
    return render(request, 'admin/fulfill_asset_modal.html', {
        'ticket': ticket,
        'available_assets': available_assets,
    })


@login_required
@require_POST
def fulfill_asset_request(request, pk):
    """Admin action to fulfill an asset request by assigning an asset."""
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    ticket = get_object_or_404(Ticket, pk=pk, status=Ticket.Status.PENDING_FULFILLMENT)
    asset_id = request.POST.get('asset_id')
    comment = request.POST.get('comment', '').strip()
    
    if not asset_id:
        messages.error(request, 'Please select an asset to assign.')
        return redirect('tickets:conversation', pk=ticket.pk)
    
    try:
        asset = Asset.objects.get(pk=asset_id)
    except Asset.DoesNotExist:
        messages.error(request, 'Asset not found.')
        return redirect('tickets:conversation', pk=ticket.pk)
    
    # Check if asset is already assigned
    if asset.assigned_to:
        messages.error(request, f'Asset {asset.name} is already assigned to {asset.assigned_to.get_full_name()}.')
        return redirect('tickets:conversation', pk=ticket.pk)
    
    # Assign asset to requester
    old_user = asset.assigned_to
    asset.assigned_to = ticket.requester
    asset.save()
    
    # Link asset to ticket
    ticket.assigned_asset = asset
    ticket.status = Ticket.Status.APPROVED
    ticket.fulfilled_at = timezone.now()
    ticket.fulfilled_by = request.user
    ticket.save()
    
    # Create asset log
    AssetLog.objects.create(
        asset=asset,
        action=AssetLog.Action.ASSIGNED,
        actor=request.user,
        details={
            'from': old_user.get_full_name() if old_user else None,
            'to': ticket.requester.get_full_name(),
            'comment': f'Fulfilled request {ticket.number}: {comment}'
        }
    )
    
    # Create ticket comment
    TicketComment.objects.create(
        ticket=ticket,
        author=request.user,
        body=f"✅ **Asset fulfilled**: {asset.name} ({asset.tracking_id}) assigned to {ticket.requester.get_full_name()}. {comment}",
        visibility='PUBLIC'
    )
    
    # Create activity log
    TicketActivityLog.objects.create(
        ticket=ticket,
        action='asset_fulfilled',
        actor=request.user,
        details={
            'asset_id': asset.pk,
            'asset_name': asset.name,
            'asset_tracking_id': asset.tracking_id,
            'assigned_to': ticket.requester.get_full_name()
        }
    )
    
    # Notify requester
    Notification.objects.create(
        recipient=ticket.requester,
        message=f'✅ Your asset request {ticket.number} has been fulfilled. {asset.name} assigned to you.',
        url=reverse('tickets:detail', args=[ticket.pk])
    )
    
    messages.success(request, f'✅ Asset {asset.name} assigned to {ticket.requester.get_full_name()}.')
    return redirect('tickets:conversation', pk=ticket.pk)


@login_required
def available_assets_for_fulfillment(request):
    """HTMX endpoint to get available assets for a specific request."""
    if request.user.role not in ['ADMIN', 'SUPERADMIN']:
        return HttpResponse(status=403)
    
    search = request.GET.get('search', '').strip()
    category = request.GET.get('category', '').strip()
    
    # Debug logging
    print(f"🔍 Search: '{search}', Category: '{category}'")
    
    # Filter available assets (unassigned and active/in-store)
    assets = Asset.objects.filter(
        assigned_to__isnull=True,
        status__in=['ACTIVE', 'IN_STORE']
    ).order_by('name')
    
    # Filter by search term
    if search:
        assets = assets.filter(
            Q(name__icontains=search) |
            Q(tracking_id__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(model__icontains=search) |
            Q(manufacturer__icontains=search)
        )
        print(f"🔍 Found {assets.count()} assets matching search")
    
    # Optional: Filter by type based on request category
    type_mapping = {
        'Hardware': ['COMPUTER', 'LAPTOP', 'PRINTER', 'SERVER', 'NETWORK'],
        'Software': ['SOFTWARE'],
        'Network': ['NETWORK', 'SERVER'],
        'Printer': ['PRINTER'],
        'Computer': ['COMPUTER', 'LAPTOP'],
    }
    
    if category in type_mapping:
        assets = assets.filter(asset_type__in=type_mapping[category])
        print(f"🔍 Filtered by category '{category}': {assets.count()} assets")
    
    # Limit results
    assets = assets[:20]
    
    return render(request, 'partials/available_assets_list.html', {
        'assets': assets,
    })